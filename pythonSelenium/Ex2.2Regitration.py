from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import random
import string

def generate_random_email():
    return ''.join(random.choice(string.ascii_lowercase) for i in range(10)) + "@yopmail.com"

def generate_random_password():
    characters = string.ascii_letters + string.digits + "!@#$%^&*()-_=+"
    return ''.join(random.choice(characters) for i in range(10))

wb = load_workbook('C:/The REST/Career/Automation Test/PyCharm_Python/Bài tập/Registration/TCs.xlsx')
ws = wb['Registration']

driver = webdriver.Chrome()

# Iterate over the steps in the Excel file
for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
    case_no, step, description = row

    register_email = ''
    register_password = ''
    test_result = ''

    try:
        if step == 1:
            driver.get('https://practice.automationtesting.in/')
            driver.maximize_window()
            test_result = 'Passed'
        elif step == 2:
            test_result = 'Passed'
        elif step == 3:
            my_account = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "a[href='https://practice.automationtesting.in/my-account/']"))
            )
            my_account.click()
            test_result = 'Passed'
        elif step == 4:
            register_email = generate_random_email()
            driver.find_element(By.CSS_SELECTOR, "input[id='reg_email']").send_keys(register_email)
            test_result = 'Passed'
        elif step == 5:
            register_password = generate_random_password()
            driver.find_element(By.CSS_SELECTOR, "input[id='reg_password']").send_keys(register_password)
            test_result = 'Passed'
        elif step == 6:
            register_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "input[class='woocommerce-Button button'][value='Register']"))
            )
            register_button.click()
            test_result = 'Passed'
        else:
            test_result = 'Failed'

        ws.cell(row=step + 1, column=5).value = test_result
        ws.cell(row=step + 1, column=6).value = register_email
        ws.cell(row=step + 1, column=8).value = register_password

    except Exception as e:
        ws.cell(row=step + 1, column=6).value = 'Failed'
        print(f"Step {step} failed with exception: {e}")

wb.save('C:/The REST/Career/Automation Test/PyCharm_Python/Bài tập/Registration/TCs.xlsx')

driver.quit()